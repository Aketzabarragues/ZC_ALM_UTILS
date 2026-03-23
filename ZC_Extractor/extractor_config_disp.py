import os
from openpyxl import load_workbook
import pandas as pd

def extraer(ruta_excel, folder_path):
    try:
        # Cargar el Excel (data_only=True para el resultado de fórmulas)
        wb = load_workbook(ruta_excel, data_only=True)
        
        nombres_a_buscar = [
            "Num_Disp_ED", 
            "Num_Disp_EA", 
            "Num_Disp_SA", 
            "Num_Disp_V", 
            "Num_Disp_M", 
            "Num_Disp_M_VF"
        ]
        
        datos_config = []

        for nombre in nombres_a_buscar:
            # Buscamos el nombre definido en el libro
            if nombre in wb.defined_names:
                # Obtenemos la ubicación (ej: 'CONFIGURACION'!$B$10)
                destinos = wb.defined_names[nombre].destinations
                
                for sheet_name, cell_address in destinos:
                    ws = wb[sheet_name]
                    valor = ws[cell_address].value
                    
                    datos_config.append({
                        "Name": nombre,
                        "Value": valor if valor is not None else 0
                    })
            else:
                print(f"   [!] Aviso: El nombre '{nombre}' no está definido en el Excel.")

        if not datos_config:
            return False, "No se encontró ningún Rango Nombrado de configuración."

        # Generar el XML como siempre
        df = pd.DataFrame(datos_config)
        output_file = os.path.join(folder_path, "config_disp.xml")
        df.to_xml(output_file, index=False, root_name="DispConfig", row_name="Item", attr_cols=False)

        return True, f"Configuración exportada vía Rangos Nombrados a {output_file}"

    except Exception as e:
        return False, f"Error: {str(e)}"